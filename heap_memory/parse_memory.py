# -*- coding: utf-8 -*-
'''按照每一行读取文件gdb.log的内容'''
from openpyxl import Workbook
from openpyxl.styles import Alignment


with open('gdb.log', 'r') as file:
    # 创建 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    addrMap = {}
    info = {}
    ws.title = "Memory Allocations"
    # 写入表头
    ws.append(["Size", "Stack Trace"])
    flag = False
    stack_trace = ""
    size = 0
    addr = None
    for line in file:
        # Process each line here
        # print(line)
        if "Breakpoint 2" in line:
            if flag:
                info["stack"] = stack_trace
                addrMap[addr] = info.copy()
                # ws.append([size, stack_trace])
            flag = False
            stack_trace = ""
            info.clear()
            size = 0
            addr = ""
            if "Breakpoint 2, zl_malloc" in line:
                flag = True
                # 获取行line中size的值
                size = line.split("size=")[1].split(')')[0]
                addr = line.split("ptr=")[1].split(',')[0]
                info["size"] = size
            elif "Breakpoint 3, zl_free" in line:
                addr = line.split("ptr=")[1].split(',')[0]
                if addr in addrMap:
                    addrMap.pop(addr)
        elif flag and '#' in line:
            if len(stack_trace) > 0:
                stack_trace += '\r\n'
            stack_trace += line.split(']')[1]
    print(addrMap)
    if flag:
        info["stack"] = stack_trace
        addrMap[addr] = info
    print(addrMap)

    # 将addrMap表里的数据写入ws中
    for addr, info in addrMap.items():
        # print(addr)
        # print(info)
        # ws.append([info["size"], info["stack"]])
        row = [info["size"], info["stack"]]
        ws.append(row)
        # 获取最后一行的单元格
        cell = ws.cell(row=ws.max_row, column=2)
        # 设置单元格的对齐方式，启用换行
        cell.alignment = Alignment(wrap_text=True)
    # 保存 Excel 文件
    wb.save("memory_allocations.xlsx")



